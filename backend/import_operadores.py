"""
Importa operadores desde Excel a PostgreSQL (tabla operadores en uph_produccion).
Ejecutar DESPUÉS de extraer_fotos.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from app.database_uph import UphSessionLocal
from app.models.uph_models import Operador

EXCEL     = Path(__file__).parent / "Listado General con fotos AMI 1.xlsx"
FOTOS_DIR = Path(__file__).parent / "uploads" / "operadores"

COL_NUM    = column_index_from_string('C') - 1  # 2
COL_NOMBRE = column_index_from_string('E') - 1  # 4

print(f"Leyendo: {EXCEL}")

wb = load_workbook(EXCEL, data_only=True)
ws = wb.active

# Misma lógica que extraer_fotos.py — escanear columna C fila por fila
empleados = {}
for row in ws.iter_rows(min_row=1):
    val = row[COL_NUM].value
    if val is None:
        continue
    num_emp = str(val).strip()
    if not num_emp.isdigit():
        continue

    fila_0 = row[0].row - 1  # 0-indexed

    # Nombre: buscar en filas cercanas columna E
    nombre = ""
    for delta in range(-3, 3):
        f = fila_0 + delta
        if f >= 0:
            c = ws.cell(row=f + 1, column=COL_NOMBRE + 1)
            if c.value:
                nombre = str(c.value).strip()
                break

    # Buscar foto existente
    foto_url = None
    for ext in ['.png', '.jpg', '.jpeg']:
        if (FOTOS_DIR / f"{num_emp}{ext}").exists():
            foto_url = f"/uploads/operadores/{num_emp}{ext}"
            break

    empleados[num_emp] = {"nombre": nombre, "foto_url": foto_url}

print(f"Empleados encontrados: {len(empleados)}")

db = UphSessionLocal()
nuevos = actualizados = sin_cambios = 0

for num_emp, data in empleados.items():
    op = db.query(Operador).filter(Operador.num_empleado == num_emp).first()
    if op:
        if op.nombre != data["nombre"] or op.foto_url != data["foto_url"]:
            op.nombre   = data["nombre"]
            op.foto_url = data["foto_url"]
            actualizados += 1
        else:
            sin_cambios += 1
    else:
        db.add(Operador(
            num_empleado=num_emp,
            nombre=data["nombre"],
            foto_url=data["foto_url"],
            activo=True,
        ))
        nuevos += 1

db.commit()
db.close()

print(f"\nResultado:")
print(f"  Nuevos:       {nuevos}")
print(f"  Actualizados: {actualizados}")
print(f"  Sin cambios:  {sin_cambios}")
print(f"  Total:        {nuevos + actualizados + sin_cambios}")
