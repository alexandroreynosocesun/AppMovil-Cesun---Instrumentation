"""
Diagnóstico v2: encuentra dónde están los datos reales
"""
import zipfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

EXCEL = Path(__file__).parent / "Listado General con fotos AMI 1.xlsx"

wb = load_workbook(EXCEL, data_only=True)
ws = wb.active

# 1. Encontrar primeras filas con datos en cualquier columna
print("=== PRIMERAS FILAS CON ALGÚN DATO (primeras 40 filas) ===")
for row in ws.iter_rows(min_row=1, max_row=40):
    valores = {get_column_letter(c.column): c.value for c in row if c.value is not None}
    if valores:
        print(f"  fila {row[0].row}: {valores}")

# 2. Buscar primera fila con número de empleado (columna C)
print("\n=== PRIMERA FILA CON DATO EN COLUMNA C ===")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    val = row[column_index_from_string('C') - 1].value
    if val is not None:
        print(f"  fila {row[0].row}: C={val}")
        # Mostrar toda esa fila
        for c in row:
            if c.value is not None:
                print(f"    col {get_column_letter(c.column)} = {c.value}")
        break

# 3. Ver distribución de imágenes por columna
print("\n=== DISTRIBUCIÓN DE IMÁGENES POR COLUMNA ===")
from collections import Counter
cols = Counter()
for img in ws._images:
    try:
        cols[img.anchor._from.col] += 1
    except:
        pass
for col_idx, count in sorted(cols.items()):
    print(f"  col {col_idx} ({get_column_letter(col_idx+1)}): {count} imágenes")

# 4. Ver imágenes en columna con más fotos (probablemente la de empleados)
print("\n=== IMÁGENES EN COLUMNA CON MÁS FOTOS (primeras 15) ===")
col_fotos = cols.most_common(1)[0][0] if cols else 18
imgs_fotos = [(img.anchor._from.row, img) for img in ws._images
              if hasattr(img.anchor, '_from') and img.anchor._from.col == col_fotos]
imgs_fotos.sort()
for fila, img in imgs_fotos[:15]:
    print(f"  fila_0idx={fila}  → fila_excel={fila+1}")
