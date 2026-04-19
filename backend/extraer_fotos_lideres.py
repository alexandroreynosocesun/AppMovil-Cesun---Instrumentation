"""
Extrae fotos del Excel de líderes AMI.
Guarda en uploads/lideres/{num_empleado}.ext
"""
import zipfile
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

EXCEL  = Path(__file__).parent / "Listado General con fotos AMI lideres .xlsx"
SALIDA = Path(__file__).parent / "uploads" / "lideres"
SALIDA.mkdir(parents=True, exist_ok=True)

# Números de empleado de los líderes (filtro)
LIDERES_NUM = {
    "202", "439", "518", "555", "2319", "7683", "9972", "10085",
    "10893", "12806", "16902", "16935", "17666", "18148",
    "19481", "20179", "26184", "33247",
}

COL_NUM    = column_index_from_string('C') - 1  # 2
COL_NOMBRE = column_index_from_string('E') - 1  # 4

NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a':   'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r':   'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
}

print(f"Leyendo: {EXCEL}\n")

# ── 1. Mapear imágenes por fila desde el drawing XML ──────────────────────────
with zipfile.ZipFile(EXCEL, 'r') as z:
    all_files = z.namelist()

    drawings  = [f for f in all_files if 'drawings/drawing' in f and f.endswith('.xml')]
    rels_files = [f for f in all_files if 'drawings/_rels' in f and f.endswith('.rels')]

    if not drawings:
        print("ERROR: No se encontró drawing XML en el Excel")
        exit(1)

    # Puede haber múltiples drawings — procesar todos
    fotos_por_fila = {}

    for drawing_file in drawings:
        # Buscar el rels correspondiente
        base = drawing_file.replace('drawings/drawing', 'drawings/_rels/drawing').replace('.xml', '.xml.rels')
        rels_file = base if base in all_files else (rels_files[0] if rels_files else None)

        rid_to_media = {}
        if rels_file and rels_file in all_files:
            tree = ET.parse(z.open(rels_file))
            for rel in tree.getroot():
                rid    = rel.attrib.get('Id', '')
                target = rel.attrib.get('Target', '')
                rid_to_media[rid] = target

        tree = ET.parse(z.open(drawing_file))
        root = tree.getroot()

        for anchor in root:
            tag = anchor.tag.split('}')[-1]
            if tag not in ('twoCellAnchor', 'oneCellAnchor'):
                continue

            from_el = anchor.find('xdr:from', NS)
            if from_el is None:
                continue
            row_el = from_el.find('xdr:row', NS)
            if row_el is None:
                continue
            row_idx = int(row_el.text)

            blip = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip')
            if blip is None:
                continue
            rid = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '')
            if not rid or rid not in rid_to_media:
                continue

            target = rid_to_media[rid]
            media_path = 'xl/' + target.replace('../', '')
            fotos_por_fila[row_idx] = media_path

print(f"Imágenes encontradas en el Excel: {len(fotos_por_fila)}")

# ── 2. Leer empleados desde la hoja ──────────────────────────────────────────
wb = load_workbook(EXCEL, data_only=True)
ws = wb.active

empleados = {}
for row in ws.iter_rows(min_row=1):
    val = row[COL_NUM].value
    if val is None:
        continue
    num_emp = str(val).strip()
    if not num_emp.isdigit():
        continue
    if num_emp not in LIDERES_NUM:
        continue  # solo procesar líderes conocidos

    fila_0 = row[0].row - 1  # 0-indexed

    nombre = ""
    for delta in range(-5, 5):
        f = fila_0 + delta
        if f >= 0:
            c = ws.cell(row=f + 1, column=COL_NOMBRE + 1)
            if c.value and str(c.value).strip():
                nombre = str(c.value).strip()
                break

    empleados[fila_0] = {"num_emp": num_emp, "nombre": nombre}

print(f"Líderes encontrados en el Excel: {len(empleados)}")

# ── 3. Extraer fotos ──────────────────────────────────────────────────────────
with zipfile.ZipFile(EXCEL, 'r') as z:
    guardados = 0
    sin_foto  = []

    for fila_0, emp in sorted(empleados.items()):
        media_path = None
        for delta in range(0, 10):
            if (fila_0 - delta) in fotos_por_fila:
                media_path = fotos_por_fila[fila_0 - delta]
                break
            if (fila_0 + delta) in fotos_por_fila:
                media_path = fotos_por_fila[fila_0 + delta]
                break

        if media_path is None:
            sin_foto.append(emp)
            continue

        try:
            ext     = Path(media_path).suffix or '.jpg'
            destino = SALIDA / f"{emp['num_emp']}{ext}"
            with z.open(media_path) as src, open(destino, 'wb') as dst:
                shutil.copyfileobj(src, dst)
            print(f"  OK  {emp['num_emp']} — {emp['nombre']}")
            guardados += 1
        except Exception as e:
            sin_foto.append(emp)
            print(f"  ERR {emp['num_emp']}: {e}")

print(f"\nFotos guardadas: {guardados}")
if sin_foto:
    print(f"Sin foto ({len(sin_foto)}):")
    for op in sin_foto:
        print(f"  {op['num_emp']} — {op['nombre']}")
