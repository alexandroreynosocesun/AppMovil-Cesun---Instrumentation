"""
Extrae fotos del Excel usando las relaciones XML internas del archivo.
Estructura del Excel (multi-fila por empleado):
  fila N:   foto en columna Q (índice 16)
  fila N+1: nombre en columna E
  fila N+2: numero de empleado en columna C
"""
import zipfile
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

EXCEL  = Path(__file__).parent / "Listado General con fotos AMI 1.xlsx"
SALIDA = Path(__file__).parent / "uploads" / "operadores"
SALIDA.mkdir(parents=True, exist_ok=True)

COL_FOTO   = 16  # columna Q, 0-indexed
COL_NUM    = column_index_from_string('C') - 1  # 2
COL_NOMBRE = column_index_from_string('E') - 1  # 4

print(f"Leyendo: {EXCEL}\n")

# ── 1. Leer posición y rId de cada imagen desde el XML del drawing ──────────
NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a':   'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r':   'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
}

with zipfile.ZipFile(EXCEL, 'r') as z:
    # Encontrar el archivo de drawing (puede ser drawing1.xml, drawing2.xml…)
    drawings = [f for f in z.namelist() if 'drawings/drawing' in f and f.endswith('.xml')]
    rels_files = [f for f in z.namelist() if 'drawings/_rels' in f and f.endswith('.rels')]

    if not drawings:
        print("ERROR: No se encontró drawing XML en el Excel")
        exit(1)

    drawing_file = drawings[0]
    rels_file    = rels_files[0] if rels_files else None

    print(f"Drawing: {drawing_file}")
    print(f"Rels:    {rels_file}\n")

    # Leer relaciones rId → target (ruta de media)
    rid_to_media = {}
    if rels_file:
        tree = ET.parse(z.open(rels_file))
        root = tree.getroot()
        for rel in root:
            rid    = rel.attrib.get('Id', '')
            target = rel.attrib.get('Target', '')
            rid_to_media[rid] = target  # ej. "../media/image5.png"

    # Leer drawing: para cada twoCellAnchor, obtener fila/col y rId
    # Estructura: <xdr:twoCellAnchor> → <xdr:from> y <xdr:pic> → <a:blipFill> → <a:blip r:embed="rIdX">
    tree = ET.parse(z.open(drawing_file))
    root = tree.getroot()

    # Imágenes de columna Q con su fila y ruta de media
    fotos_por_fila = {}  # fila_0idx → ruta_en_zip

    for anchor in root:
        tag = anchor.tag.split('}')[-1]
        if tag not in ('twoCellAnchor', 'oneCellAnchor'):
            continue

        # Obtener posición _from
        from_el = anchor.find('xdr:from', NS)
        if from_el is None:
            continue
        col_el = from_el.find('xdr:col', NS)
        row_el = from_el.find('xdr:row', NS)
        if col_el is None or row_el is None:
            continue
        col_idx = int(col_el.text)
        row_idx = int(row_el.text)  # 0-indexed

        if col_idx != COL_FOTO:
            continue  # solo columna Q

        # Obtener rId del blip
        blip = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip')
        if blip is None:
            continue
        rid = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '')
        if not rid or rid not in rid_to_media:
            continue

        # Normalizar ruta: "../media/imageX.png" → "xl/media/imageX.png"
        target = rid_to_media[rid]
        media_path = 'xl/' + target.replace('../', '')

        fotos_por_fila[row_idx] = media_path

print(f"Fotos mapeadas en columna Q: {len(fotos_por_fila)}")

# ── 2. Leer empleados desde la hoja ──────────────────────────────────────────
wb = load_workbook(EXCEL, data_only=True)
ws = wb.active

empleados = {}
for row in ws.iter_rows(min_row=1):
    val = row[COL_NUM].value
    if val is None:
        continue
    num_emp = str(val).strip()
    if not num_emp.isdigit():
        continue

    fila_0 = row[0].row - 1  # 0-indexed

    # Nombre: buscar en filas cercanas columna E
    nombre = ""
    for delta in range(-3, 3):
        f = fila_0 + delta
        if f >= 0:
            c = ws.cell(row=f + 1, column=COL_NOMBRE + 1)
            if c.value:
                nombre = str(c.value).strip()
                break

    empleados[fila_0] = {"num_emp": num_emp, "nombre": nombre}

print(f"Empleados encontrados: {len(empleados)}")

# ── 3. Extraer fotos ──────────────────────────────────────────────────────────
with zipfile.ZipFile(EXCEL, 'r') as z:
    guardados = 0
    sin_foto  = []

    for fila_0, emp in sorted(empleados.items()):
        # Buscar foto en filas cercanas (foto suele estar 2 filas antes del num_emp)
        media_path = None
        for delta in range(0, 6):
            fila_buscar = fila_0 - delta
            if fila_buscar in fotos_por_fila:
                media_path = fotos_por_fila[fila_buscar]
                break

        if media_path is None:
            sin_foto.append(emp)
            continue

        try:
            ext     = Path(media_path).suffix
            destino = SALIDA / f"{emp['num_emp']}{ext}"
            with z.open(media_path) as src, open(destino, 'wb') as dst:
                shutil.copyfileobj(src, dst)
            print(f"  OK  {emp['num_emp']} — {emp['nombre']}")
            guardados += 1
        except Exception as e:
            sin_foto.append(emp)
            print(f"  ERR {emp['num_emp']}: {e}")

print(f"\nFotos guardadas: {guardados}")
if sin_foto:
    print(f"Sin foto ({len(sin_foto)}):")
    for op in sin_foto[:20]:
        print(f"  {op['num_emp']} — {op['nombre']}")
    if len(sin_foto) > 20:
        print(f"  ... y {len(sin_foto) - 20} más")
